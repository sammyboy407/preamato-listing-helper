"""Assembles listing rows in the exact column layout of the given eBay
template.

The template itself (its Categories/Aspects/ConditionDescriptors/
BusinessPolicy sheets) is still the source of truth for which categories,
conditions, and item-specific values are valid — see ebay_template.py and
content_generator.py. The delivered output, however, is a flat CSV matching
the template's Listings sheet's exact header row, since that's what this
account's upload flow actually accepts.

The auto-generated template only pre-populates a subset of the columns
eBay's uploader actually recognises for a given category selection — it
doesn't include every field just because a category could theoretically use
it. Confirmed against a real historical upload file: OriginalRetailPrice
(the strikethrough "was" price), ConditionDescription (a dedicated field,
distinct from the free-text Description), and C:Country of Origin (present
in this account's aspects for every tested category, but excluded from the
template's own header row) are all legitimate columns eBay accepts even
though the downloaded template omits them — see EXTRA_COLUMNS.
"""
from __future__ import annotations

import csv
import math
import re
import warnings
import zipfile
from pathlib import Path

# Recognised eBay fields to append after the template's own header row, even
# when the specific template download didn't include them (see module
# docstring). C:Country of Origin's value is already computed by
# content_generator.py whenever a category's Aspects define it — it was
# just being silently dropped by write_csv's header-restricted output
# before this was added.
EXTRA_COLUMNS = ["OriginalRetailPrice", "ConditionDescription", "C:Country of Origin"]

from . import config, ebay_template
from .data_loader import Product, split_image_urls

GENDER_POSSESSIVE_MAP = {"WOMEN": "Women's", "MEN": "Men's", "UNISEX": "Unisex"}


def _gender_possessive(raw) -> str:
    if not raw:
        return ""
    key = str(raw).strip().upper()
    return GENDER_POSSESSIVE_MAP.get(key, f"{str(raw).strip().title()}'s")


def _format_rrp(rrp) -> str:
    try:
        return f"£{float(rrp):.2f}"
    except (TypeError, ValueError):
        return "N/A"


def _primary_colour_key(category_id: str, template: ebay_template.EbayTemplate) -> str | None:
    aspects = template.aspects.get(str(category_id), {})
    colour_names = [n for n in aspects if "colour" in n.lower()]
    required = [n for n in colour_names if aspects[n].level == "REQUIRED"]
    if required:
        return required[0]
    return colour_names[0] if colour_names else None


def build_description(
    product: Product,
    ai_result: dict,
    category: ebay_template.CategorySpec,
    template: ebay_template.EbayTemplate,
    blurb_cache: dict[str, str],
) -> str:
    m, meas = product.master, product.measurements
    brand = m.get("Brand") or ""

    paragraph = blurb_cache.get(brand, "")

    condition_notes = (meas.get("Description") or "").strip()
    condition_line = (
        ai_result.get("condition_description") if condition_notes
        else config.CONDITION_PLACEHOLDER
    )

    specifics = ai_result.get("item_specifics", {})
    colour_key = _primary_colour_key(category.category_id, template)
    colour = specifics.get(colour_key, "") if colour_key else ""
    style = specifics.get("C:Style", "Not Specified")
    size = specifics.get("C:Size") or specifics.get("C:UK Shoe Size") or meas.get("Size") or m.get("Size") or ""

    department_line = ", ".join(
        part for part in (
            _gender_possessive(m.get("Gender")),
            m.get("Category"),
            m.get("Season"),
        ) if part
    )

    # Trailing space on each spec line, matching the exact format requested.
    spec_lines = [
        f"Condition: {condition_line}",
        f"Brand: {brand}",
        f"Style: {style}",
        f"Colour: {colour}",
        f"Material: {ai_result.get('material_summary', 'Not Specified')}",
        f"Size: {size}",
        f"Department: {department_line}",
        f"RRP: {_format_rrp(m.get('Rounded RRP'))}",
    ]

    lines = [paragraph, ""] + [line + " " for line in spec_lines] + [config.SHIPPING_LINE]
    # eBay renders Description as HTML — a plain "\n" is whitespace that
    # gets collapsed on the live listing page, so line breaks only actually
    # show up in the CSV/spreadsheet, never on eBay itself. <br> forces a
    # real visible break there; keeping "\n" alongside it just makes the
    # raw text still readable if someone opens the file directly.
    return "<br>\n".join(lines)


def compute_start_price(rrp, price_percent: float) -> float | None:
    """Selling price as a % of RRP, rounded to the nearest 5 (whole currency
    units — e.g. 41.50 -> 40, 42.50 -> 45). Plain round() uses banker's
    rounding (round-half-to-even), which would round exact halfway points
    inconsistently for money — this always rounds halfway points up."""
    try:
        raw = float(rrp) * (price_percent / 100)
    except (TypeError, ValueError):
        return None
    return 5 * math.floor(raw / 5 + 0.5)


def build_row(
    product: Product,
    ai_result: dict,
    category: ebay_template.CategorySpec,
    template: ebay_template.EbayTemplate,
    blurb_cache: dict[str, str],
    schedule_time: str | None = None,
    price_percent: float = config.START_PRICE_RATIO * 100,
) -> dict:
    m, meas = product.master, product.measurements

    rrp = m.get("Rounded RRP") or 0
    start_price = compute_start_price(rrp, price_percent)

    pic_urls = split_image_urls(meas.get("Images 2D link"))

    row = {
        "*Action(SiteID=UK|Country=GB|Currency=GBP|Version=1193)": config.ACTION,
        "Custom label (SKU)": product.sku,
        "Category ID": category.category_id,
        "Category name": category.category_name,
        "Title": ai_result.get("title"),
        "Start price": start_price,
        "OriginalRetailPrice": rrp,
        "Quantity": m.get("Qty") or config.QUANTITY_DEFAULT,
        "Item photo URL": "|".join(pic_urls),
        "Condition ID": ai_result.get("condition_id"),
        "ConditionDescription": ai_result.get("condition_description"),
        "Description": build_description(product, ai_result, category, template, blurb_cache),
        "Format": config.FORMAT,
        "Duration": config.DURATION,
        "Best Offer Enabled": config.BEST_OFFER_ENABLED,
        "VAT%": config.VAT_PERCENT,
        "Immediate pay required": True,
        "Location": config.LOCATION,
        "Shipping profile name": config.SHIPPING_PROFILE,
        "Return profile name": config.RETURN_PROFILE,
        "Payment profile name": config.PAYMENT_PROFILE,
    }

    for field_name, value in ai_result.get("item_specifics", {}).items():
        row[field_name] = value

    if schedule_time and "Schedule Time" in template.listing_headers:
        row["Schedule Time"] = schedule_time

    return row


def _patch_family_values(xml_bytes: bytes) -> bytes:
    """Fixes a font `family` value >14 that openpyxl's schema rejects on
    load — seen in some eBay-exported template files."""
    text = xml_bytes.decode("utf-8", errors="ignore")
    text = re.sub(r'family val="(\d+)"', lambda m: 'family val="2"' if int(m.group(1)) > 14 else m.group(0), text)
    return text.encode("utf-8")


def _patch_template_bytes(template_path: str | Path, patched_path: str | Path) -> None:
    with zipfile.ZipFile(template_path, "r") as zin, zipfile.ZipFile(patched_path, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename.endswith(".xml") and b"family val=" in data:
                data = _patch_family_values(data)
            zout.writestr(item, data)


def write_into_template(rows: list[dict], template_path: str | Path, output_path: str | Path) -> None:
    import openpyxl  # local import: only needed for this write step

    patched_path = Path(output_path).with_suffix(".patched_tmp.xlsx")
    _patch_template_bytes(template_path, patched_path)

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")  # "Data Validation extension is not supported"
        wb = openpyxl.load_workbook(patched_path)

    ws = wb["Listings"]
    headers = [c.value for c in ws[4]]  # row 4 is the real header row (rows 1-3 are #INFO)

    for row_offset, row in enumerate(rows):
        excel_row = 5 + row_offset
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=excel_row, column=col_idx, value=row.get(header))

    wb.save(output_path)
    patched_path.unlink(missing_ok=True)


def _csv_cell(value):
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    return value


def output_headers(template: ebay_template.EbayTemplate) -> list[str]:
    """The template's own header row plus any EXTRA_COLUMNS it didn't
    already include."""
    return template.listing_headers + [c for c in EXTRA_COLUMNS if c not in template.listing_headers]


def write_csv(rows: list[dict], template: ebay_template.EbayTemplate, output_path: str | Path) -> None:
    """Writes a flat CSV matching the template's Listings sheet, plus any
    EXTRA_COLUMNS appended at the end: its #INFO preamble rows first (eBay's
    uploader uses these to identify the template — a CSV missing them is
    rejected with "We couldn't identify your template"), padded out to the
    extended width so the file stays rectangular, then the (extended)
    header row, then data. No other sheets. utf-8-sig so Excel (incl. on
    Windows) auto-detects UTF-8 and renders the £ symbol correctly."""
    headers = output_headers(template)
    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        for info_row in template.info_rows:
            writer.writerow(info_row + [""] * (len(headers) - len(info_row)))
        writer.writerow(headers)
        for row in rows:
            writer.writerow([_csv_cell(row.get(h)) for h in headers])
